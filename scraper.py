# scraper.py
import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from services.process import Process
import time
from selenium.webdriver.common.by import By
import openpyxl
from tkinter import *
from tkinter import ttk

# GUI設定
root = Tk()
root.title('領事館スクレイピングツール')
frame = ttk.Frame(root, padding=16)

root.mainloop()

def main():
    BrowserPath=ResourcePath("./browser/chrome.exe") # ブラウザ
    DriverPath=ResourcePath("./driver/chromedriver.exe") # ウェブドライバ

    # ウェブドライバ設定
    options=Options()
    options.binary_location=BrowserPath
    # options.add_argument("--headless") # 動きを見たい場合はコメントアウトする。
    driver=webdriver.Chrome(DriverPath, options=options)

    # スクレイピング
    ProcessC=Process(driver)
    ProcessC.goPage()
    time.sleep(10)

    # 変数宣言
    embassy_names = []
    embassy_names_eng = []
    post_codes = []
    address_list = []
    tel_codes = []
    areas = []

    # 大使館＆領事館名取得
    embassy_names = driver.find_elements(By.TAG_NAME, "h3")
    # 大使館＆領事館名(英語)取得
    embassy_names_eng = driver.find_elements(By.TAG_NAME, "h4")
    # 住所&郵便番号取得
    post_codes = driver.find_elements(By.XPATH, "//div/div/div/div/div/p")
    # 電話番号
    tel_codes = driver.find_elements(By.XPATH, "//div/div/div/div/div/p")
    # 管理区域
    areas = driver.find_elements(By.XPATH, "//div/div/div/div/div/p")

    # エクセル作成
    path = r'C:/Users/uu/Desktop/Embassy_List.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb["Europe"]

    # 大使館＆領事館名
    RowNumber = 2
    for text in embassy_names:
        if text != '\n\n\n\n':
            ws.cell(RowNumber, 2).value = text.text
            RowNumber = RowNumber + 1
    # 大使館＆領事館名(英語)
    RowNumber = 2
    for text in embassy_names_eng:
        if text != '\n\n\n\n':
            ws.cell(RowNumber, 3).value = text.text
            RowNumber = RowNumber + 1
    # 郵便番号
    RowNumber = 2
    for text in post_codes:
        if text != '\n\n\n\n':
            if text.text != '駐日外国公館リストの目次へ戻る':
                ws.cell(RowNumber, 4).value = text.text[0:9]
                RowNumber = RowNumber + 1
    # 住所
    RowNumber = 2
    for text in post_codes:
        telidx = text.text.find('電話')
        if text != '\n\n\n\n' and not text.text.__contains__('駐日外国公館リストの目次へ戻る'):
            if text.text != '駐日外国公館リストの目次へ戻る':
                ws.cell(RowNumber, 5).value = text.text[10:telidx - 1]
                RowNumber = RowNumber + 1

    # 電話番号
    RowNumber = 2
    for text in tel_codes:
        idx = 0
        if text != '\n\n\n\n' and not text.text.__contains__(
                '駐日外国公館リストの目次へ戻る') and text.text != '駐日外国公館リストの目次へ戻る':
            idx = text.text.find('電話')
            if idx != -1:
                if text.text[idx + 15] == '、':
                    ws.cell(RowNumber, 6).value = text.text[idx + 3:idx + 28]
                else:
                    ws.cell(RowNumber, 6).value = text.text[idx + 3:idx + 15]
            else:
                ws.cell(RowNumber, 6).value = text.text

        RowNumber = RowNumber + 1
    # 管轄区域
    RowNumber = 2
    for text in tel_codes:
        idx = 0
        if text != '\n\n\n\n' and not text.text.__contains__(
                '駐日外国公館リストの目次へ戻る') and text.text != '駐日外国公館リストの目次へ戻る':
            idx = text.text.find('管轄区域')
            if idx != -1:
                ws.cell(RowNumber, 7).value = text.text[idx + 5:]
            else:
                idx2 = text.text.find('管轄')
                if idx2 != -1:
                    ws.cell(RowNumber, 7).value = text.text[idx + 3:]
                else:
                    ws.cell(RowNumber, 7).value = ''
            RowNumber = RowNumber + 1
    wb.save(path)
    wb.close()

    # クローズ処理
    time.sleep(10)
    driver.close()
    driver.quit()

def ResourcePath(relativePath):
    try:
        basePath=sys._MEIPASS
    except Exception:
        basePath=os.path.dirname(__file__)
    return os.path.join(basePath, relativePath)

if __name__=="__main__":
    main()
